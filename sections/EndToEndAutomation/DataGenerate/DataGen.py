import openpyxl

def dataGenerator():
    wk = openpyxl.load_workbook('../list.xlsx')
    sh = wk['Sheet1']
    r = sh.max_row
    li = [[(sh.cell(i,1)).value,(sh.cell(i,2)).value] for i in range(1,r+1)]
    print(li)

    return li